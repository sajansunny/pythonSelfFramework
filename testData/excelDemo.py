import openpyxl
dict1 = {}
book = openpyxl.load_workbook("C:\\Users\\admin\\Documents\\TestData\\HomePageTestData.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

# sheet.cell(row=4, column=2).value = "Sunny"
# print(sheet.cell(row=4, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['D3'].value)

for i in range(1, sheet.max_row+1):
    # if sheet.cell(row=i, column=1).value == "TC2":
        for j in range(2, sheet.max_column+1):
            # print(sheet.cell(row=i, column=j).value)
            dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict1)
