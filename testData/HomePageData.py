import openpyxl


class HomePageData:
    test_HomePage_data = [{"name": "Sajan", "email": "sajan@mail.com", "gender": "Male"}, {"name": "Shelmi", "email": "shelmi@mail.com", "gender": "Female"}]

    @staticmethod
    def getTestData(TC_id):
        book = openpyxl.load_workbook("C:\\Users\\admin\\Documents\\TestData\\HomePageTestData.xlsx")
        # book = openpyxl.load_workbook('.\\testData\\HomePageTestData.xlsx')
        sheet = book.active
        dict1 = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == TC_id:
                for j in range(2, sheet.max_column + 1):
                    dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dict1]
